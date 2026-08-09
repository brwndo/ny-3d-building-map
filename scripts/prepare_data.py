"""One-time data prep: flattens the ESG spreadsheet into public/data/buildings.geojson.

Reads three sheets from `Enertiv Product Management/25-asset-level-esg-data-NY.xlsx`:
  - Asset-Level Data          (one row per asset)
  - Baseline & Targets        (long format: one row per asset per metric, pivoted here)
  - Compliance Exposure (LL97) (one row per asset; ignores the side lookup table in cols U-W)

Also converts the Census NY state boundary shapefile to GeoJSON if present at
scripts/source/cb_2023_us_state_500k.shp (see prepare instructions in README).

Run:  .venv/bin/python scripts/prepare_data.py
Re-run manually whenever the source xlsx changes. Not part of the running app.
"""

import json
import math
import os
from datetime import date, datetime, timedelta

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "Enertiv Product Management", "25-asset-level-esg-data-NY.xlsx")
OUT_DIR = os.path.join(ROOT, "public", "data")
BOUNDARY_SHP = os.path.join(ROOT, "scripts", "source", "cb_2023_us_state_500k.shp")

# Metric display strings in "Baseline & Targets" col C -> internal keys
METRIC_KEYS = {
    "GHG Emissions (Scope 1+2, location-based)": "ghg",
    "Energy Use Intensity (EUI)": "eui",
    "Water Use (WT1)": "water",
    "Waste Diversion Rate (WS1)": "waste",
    "ENERGY STAR Score": "energyStar",
    "Data Coverage": "coverage",
}
# Fixed-goal metrics color by distance to the single portfolio-wide target;
# the sheet's % Progress column explodes for these when baseline is near target.
FIXED_GOAL = {"energyStar", "waste", "coverage"}

# Thresholds derived empirically from the sheet's own Status column
# (Target Met >= 1.0, On Track >= 0.75, At Risk >= 0.40, else Off Track) so
# trajectory-metric bands agree with the sheet exactly. Fixed-goal metrics
# use our distance-to-target p (the sheet's % Progress explodes for them),
# so residual disagreement with the sheet's Status there is expected and
# intentional - the computed band is the displayed status for those.
BAND_THRESHOLDS = [
    (1.0 - 1e-9, "Target Met"),
    (0.75, "On Track"),
    (0.40, "At Risk"),
]
EPS = 1e-12


def band_for(p):
    for threshold, label in BAND_THRESHOLDS:
        if p >= threshold:
            return label
    return "Off Track"


def clamp01(x):
    return max(0.0, min(1.0, x))


def header_map(ws):
    """Header text -> 1-based column index, first occurrence wins (the
    Compliance sheet repeats 'Property Type' in its side lookup table)."""
    headers = {}
    for cell in ws[1]:
        if cell.value is not None and cell.value not in headers:
            headers[str(cell.value).strip()] = cell.column
    return headers


def sheet_rows(ws, key_col="Asset ID"):
    """Yield dict rows keyed by header, skipping rows with an empty Asset ID
    (row 32 is blank in both per-asset sheets)."""
    headers = header_map(ws)
    key_idx = headers[key_col]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if key_idx - 1 >= len(row) or row[key_idx - 1] in (None, ""):
            continue
        yield {h: (row[i - 1] if i - 1 < len(row) else None) for h, i in headers.items()}


def as_date_str(value):
    """Normalize the sheet's mixed date typing: text 'YYYY-MM-DD', datetime
    objects, or raw Excel serial numbers."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
    return str(value).strip()


def none_str(value):
    """Certification columns use the literal string 'None' for no certification."""
    if value is None:
        return None
    text = str(value).strip()
    return None if text == "None" or text == "" else text


def compute_metric(row):
    """Returns the per-metric record incl. the normalized, direction-adjusted,
    clamped color value p and its band label."""
    direction = row["Direction"]
    current = float(row["Current Value"])
    target = float(row["Target Value"])
    progress = float(row["% Progress to Target"])
    key = METRIC_KEYS[row["Metric"]]

    if key in FIXED_GOAL:
        if direction == "Higher is better":
            p = clamp01(current / target) if abs(target) > EPS else 1.0
        else:
            p = clamp01(target / current) if abs(current) > EPS else 1.0
    else:
        p = clamp01(progress)

    return key, {
        "unit": row["Unit"],
        "direction": direction,
        "baselineValue": row["Baseline Value"],
        "baselineYear": row["Baseline Year"],
        "currentValue": current,
        "targetValue": target,
        "targetYear": row["Target Year"],
        "progressPct": progress,
        "varianceVsTarget": row["Variance vs Target"],
        "sheetStatus": row["Status"],
        "p": round(p, 6),
        "band": band_for(p),
    }


def build_features(wb):
    assets = list(sheet_rows(wb["Asset-Level Data"]))
    baselines = list(sheet_rows(wb["Baseline & Targets"]))
    compliance = {r["Asset ID"]: r for r in sheet_rows(wb["Compliance Exposure (LL97)"])}

    metrics_by_asset = {}
    for row in baselines:
        key, record = compute_metric(row)
        metrics_by_asset.setdefault(row["Asset ID"], {})[key] = record

    features = []
    for a in assets:
        asset_id = a["Asset ID"]
        c = compliance[asset_id]
        score = int(a["ENERGY STAR Score (1-100)"])
        props = {
            "id": asset_id,
            "name": a["Asset Name"],
            "floors": int(a["Number of Floors"]),
            "address": a["Address"],
            "city": a["City"],
            "state": a["State"],
            "zip": str(a["Zip Code"]),
            "propertyType": a["Property Type"],
            "floorArea": float(a["Floor Area (sq ft)"]),
            "raw": {
                "en1Imported": a["EN1 - Energy Imported (kWh)"],
                "en1Generated": a["EN1 - Energy Generated On-site (kWh)"],
                "en1Exported": a["EN1 - Energy Exported (kWh)"],
                "eui": a["EUI (kWh/sq ft)"],
                "scope1": a["GH1 - Scope 1 (mtCO2e)"],
                "scope2Location": a["GH1 - Scope 2 Location-based (mtCO2e)"],
                "scope2Market": a["GH1 - Scope 2 Market-based (mtCO2e)"],
                "scope3": a["GH1 - Scope 3, unscored (mtCO2e)"],
                "wt1": a["WT1 - Water Withdrawals (gal)"],
                "wt2": a["WT2 - Discharge to Sensitive Waterways (gal)"],
                "ws1": a["WS1 - Waste Diversion Rate (%)"],
            },
            "cert": {
                "bc11": none_str(a["BC1.1 - Design/Construction Certification"]),
                "bc12": none_str(a["BC1.2 - Operational Certification"]),
                "bc2": none_str(a["BC2 - Ongoing Certification"]),
            },
            "energyStarScore": score,
            "energyStarEligible": score >= 75,
            "dataCoverage": float(a["Data Coverage (%)"]),
            "confidenceTier": a["Confidence Tier"],
            "dataSource": a["Data Source"],
            "lastUpdated": as_date_str(a["Last Updated"]),
            "metrics": metrics_by_asset[asset_id],
            "compliance": {
                "ll97Applicable": c["LL97 Applicable"],
                "currentEmissions": c["Current Annual Emissions - Scope 1+2 (mtCO2e)"],
                "limit2024": c["2024-2029 Emissions Limit (mtCO2e)"],
                "overCap2024": c["Over/(Under) Cap 2024-2029 (mtCO2e)"],
                "fine2024": c["Estimated Fine Exposure 2024-2029 ($)"],
                "limit2030": c["2030-2034 Emissions Limit (mtCO2e)"],
                "overCap2030": c["Projected Over/(Under) Cap 2030-2034 (mtCO2e)"],
                "fine2030": c["Projected Fine Exposure 2030-2034 ($)"],
                "ll84FilingStatus": c["LL84 Filing Status"],
                "ll84NextDeadline": as_date_str(c["LL84 Next Filing Deadline"]),
                "daysToDeadline": c["Days to Deadline"],
                "nonFilingPenaltyRate": c["Non-Filing Penalty Rate ($/month if unfiled)"],
                "otherLaws": c["Other NYC Local Laws Applicable"],
                "status2024": c["Compliance Status 2024-2029"],
                "status2030": c["Compliance Status 2030-2034 (Projected)"],
            },
        }
        features.append({
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [float(a["Longitude"]), float(a["Latitude"])],
            },
            "properties": props,
        })
    return features


def verify_bands(features):
    """Compare our computed band with the sheet's precomputed Status per
    asset per metric. Prints an agreement summary and every mismatch."""
    total, matches, mismatches = 0, 0, []
    for f in features:
        for key, m in f["properties"]["metrics"].items():
            total += 1
            if m["band"] == m["sheetStatus"]:
                matches += 1
            else:
                mismatches.append((f["properties"]["id"], key, m["p"], m["band"], m["sheetStatus"]))
    print(f"Band verification: {matches}/{total} agree with sheet Status")
    for asset_id, key, p, band, status in mismatches:
        print(f"  MISMATCH {asset_id} {key}: p={p:.3f} computed={band!r} sheet={status!r}")


def convert_boundary():
    if not os.path.exists(BOUNDARY_SHP):
        print(f"Boundary shapefile not found at {BOUNDARY_SHP} - skipping boundary conversion")
        return
    import shapefile

    reader = shapefile.Reader(BOUNDARY_SHP)
    for sr in reader.shapeRecords():
        if sr.record["STATEFP"] == "36":
            geo = {
                "type": "FeatureCollection",
                "features": [{
                    "type": "Feature",
                    "geometry": sr.shape.__geo_interface__,
                    "properties": {"name": sr.record["NAME"], "statefp": "36"},
                }],
            }
            out = os.path.join(OUT_DIR, "ny-state-boundary.geojson")
            with open(out, "w") as fh:
                json.dump(geo, fh)
            print(f"Wrote {out}")
            return
    print("STATEFP 36 not found in shapefile")


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    features = build_features(wb)
    print(f"Built {len(features)} features")

    verify_bands(features)

    out = os.path.join(OUT_DIR, "buildings.geojson")
    with open(out, "w") as fh:
        json.dump({"type": "FeatureCollection", "features": features}, fh)
    print(f"Wrote {out}")

    convert_boundary()


if __name__ == "__main__":
    main()
